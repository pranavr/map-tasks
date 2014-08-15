import json

import datetime as datetime
import clusters
import urllib
import urllib2
import time
import requests
import utils as utils
import geojson as geojson
import xml.etree.cElementTree as et
import openpyxl as xlreader
def readXLSX(place):
	filename = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/2010 to 2012 all Accident Data.xlsx'
	workbook = xlreader.load_workbook(filename)
	sheet = workbook.get_active_sheet()
	row = 2
	jsonData = [[], [], []]
	streetData = {'2010':{}, '2011':{}, '2012':{}}
	while (row < 10139):
		lat = sheet.cell('D' + str(row)).value	
		lng = sheet.cell('E' + str(row)).value
		city = str(sheet.cell('I' + str(row)).value)

		if ((city.upper() == place.upper())) and ((lat != '') and (lng != '')):
			lng = float(lng) * -1
			jsonObj = {}
			jsonObj['lat'] = float(lat)
			jsonObj['lng'] = float(lng)
			year = str(sheet.cell('A' + str(row)).value)
			jsonObj['year'] = year
			jsonObj['date'] = str(sheet.cell('F' + str(row)).value)
			jsonObj['severity'] = sheet.cell('M' + str(row)).value
			street1 = str(sheet.cell('O' + str(row)).value).strip()
			street2 = str(sheet.cell('P' + str(row)).value).strip()
			

			jsonObj['street1'] = street1
			jsonObj['street2'] = street2
			jsonData[int(year[3])].append(jsonObj)


			if not (street1 == '(N/A)' or street1 == None or street1 =='(UNK)'):
				if street1 not in streetData[year].keys():
					streetData[year][street1] = [[lat, lng]]
				else:
					streetData[year][street1].append([lat, lng])

			if not (street2 == '(N/A)' or street2 == None or street2 == '(UNK)'):
				if street2 not in streetData[year].keys():
					streetData[year][street2] = [[lat, lng]]
				else:
					streetData[year][street2].append([lat, lng])

		
		row+=1
	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/final_data.json', 'w')
	finalJSON= {'city': place, 'accidents':{'2010':jsonData[0], '2011':jsonData[1], '2012': jsonData[2]}, 'clusters': streetData}
	json.dump(finalJSON, outputfile)
	print len(jsonData[0]), len(jsonData[2]), len(jsonData[2])


def mergeJSONdata(place):
	jsonFile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/bicycle_accidents/final_data_streets.json', 'r')
	jsonData = json.load(jsonFile)
	streetData = jsonData['clusters']
	accidents = jsonData['accidents']
	merged_accidents = []
	merged_clusters = []
	
	keys = accidents.keys()
	for i in range(len(keys)):
		for acc in accidents[keys[i]]:
			merged_accidents.append(acc)

	#street_years = streetData.keys()
	merged_streets = streetData[keys[0]]

	for j in range(1,len(keys)):
		yearData = streetData[keys[j]]
		for street in yearData.keys():
			if street in merged_streets.keys():
				for yds in yearData[street]:
					merged_streets[street].append(yds)
			else:
				merged_streets[street] = yearData[street]

	for st in merged_streets.keys():
		streets_points = merged_streets[st]
		street_clusters = clusters.createClusters(streets_points, 0.02)
		if len(street_clusters) > 0:
			for st_c in street_clusters:
				merged_clusters.append(st_c);

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/final_data2.json', 'w')
	finalJSON = {'city':place, 'accidents':merged_accidents, 'streets':merged_streets, 'clusters':merged_clusters}
	json.dump(finalJSON, outputfile)



# If city has data for more than one year (good amount)
def makeStreetClusters(place):
	jsonFile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/bicycle_accidents/final_data_2013.json', 'r')
	jsonData = json.load(jsonFile)
	accidents = jsonData['accidents']
	streetData = jsonData['streets']
	clusters_data = []
	# for year in streetData.keys():
	# 	clusters_data[year] = []
	for street in streetData.keys():

		streets_points = streetData[street]
		print streets_points
		street_clusters = clusters.createClusters(streets_points, 0.015)
		if len(street_clusters) > 0:
			for st_c in street_clusters:
				clusters_data.append(st_c)

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/bicycle_accidents/final_corrected_data_2013.json', 'w')
	finalJSON = {'city':place, 'accidents':accidents, 'streets':streetData, 'clusters':clusters_data}
	json.dump(finalJSON, outputfile)


def downloadImage(lattitude, longitude,heading,pitch, outputfile, api_key='AIzaSyBZnvqy9HEpG-LAQwm_AxDOegMciI9jgP4'):
	url = 'http://maps.googleapis.com/maps/api/streetview?size=640x640&location='+str(lattitude)+','+str(longitude)+'&fov=120&heading='+str(heading)+'&pitch='+str(pitch)+'&sensor=false&key='+api_key
	response = requests.get(url)
	localfile = open(str(outputfile), 'w')
   	
   	localfile.write(response.content)
   	localfile.close()


def downloadStreetViews(place,i):
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_corrected_data_2013.json', 'r'))
	accidents = jsonData['accidents']
	streetData = jsonData['streets']
	clusters_data = jsonData['clusters']
	accidents_with_images = []
	outputfilename = basePath + 'streetviews/'
	while  i < len(accidents):
		acc  = accidents[i]
		lat = acc['lat']
		lng = acc['lng']
		acc['image_id'] = i
		filename = outputfilename + place + '_' + str(i) + '.jpeg'
		downloadImage(lat, lng, 0, 0, filename)
		print 'done with ' + str(i) + ' out of ' + str(len(accidents)-1)
		i +=1
		accidents_with_images.append(acc)
		time.sleep(0.3)


	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(place) + '/bicycle_accidents/final_data_2013.json', 'w')
	finalJSON = {'city':place, 'accidents':accidents_with_images, 'streets':streetData, 'clusters':clusters_data}
	json.dump(finalJSON, outputfile)


def get_street_name(lat,lng):
	url = 'http://maps.googleapis.com/maps/api/geocode/xml?latlng='+str(lat)+','+str(lng)+'&sensor=false'
	url2 = 'http://api.geonames.org/findNearestIntersectionOSM?lat='+str(lat)+'&lng='+str(lng)+'&username=pranav'
	#response = requests.get(utils.createSecureURL(url))
	print url2
	response = requests.get(url2)
	parse = et.XML(response.content)

	if parse != None and parse.find('intersection') != None:
		street1_name = parse.find('intersection').find('street1').text
		street2_name = parse.find('intersection').find('street2').text

		print street1_name, street2_name

		# if len(street1_names)>= 2:
		# 	s1_name = street1_names[1] + ' ' + street1_names[2]
		# else:
		# 	s1_name = street1_names[1]

		# if len(street2_names)>= 2:
		# 	s2_name = street2_names[1] + ' ' + street2_names[2]
		# else:
		# 	s2_name = street2_names[1]

		street1_name.replace(u'\xe9', 'e')
		street2_name.replace(u'\xe9', 'e')
		#print s1_name,s2_name
		return [street1_name, street2_name]
	else:
		time.sleep(0.2)
		print 'error'	
	# if parse.find('status').text =='OK':
	# 	address_components = parse.find('result').findall('address_component')
	# 	print address_components

	# 	for address_component in address_components:
	# 		long_name = address_component.find('long_name').text
	# 		type_name = address_component.find('type').text
	# 		if type_name == 'route':
	# 			names = long_name.split()
	# 			street_name = str(names[1]) +' '+ str(names[2])
	# 			return street_name
	# else:
	# 	print "error"
	# 	return None


def portland_street():
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Portland/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data_streets.json', 'r'))
	outCSV = open(basePath+'/stretnames2.csv', 'r')
	streetNames= outCSV.readlines()
	accidents_streets = {}
	for line in streetNames:
		line = line.split(',')
		image_id = int(line[0])
		street1 = str(line[1]).strip()
		street2 = str(line[2][:len(line[2])-1]).strip()
		accidents_streets[image_id] = [street1, street2]


	accidents = jsonData['accidents']
	streetData = {}
	clusters_data = jsonData['clusters']
	accidents_with_streetnames = []
	i = 0
	while (i < len(accidents)):
		acc  = accidents[i]
		lat = acc['lat']
		lng = acc['lng']


		[street1, street2] = accidents_streets[int(acc['image_id'])]
		acc['street1'] = street1.strip()
		acc['street2'] = street2.strip()

		#outCSV.write(str(lat) +','+ str(lng) +','+street1_name + ',' +street2_name+'\n')
		i+=1
		accidents_with_streetnames.append(acc)
		print 'done with ' + str(i) + 'out of ' + str(len(accidents))

		if not (street1 == '(N/A)' or street1 == None or street1 =='(UNK)'):
			if street1 not in streetData.keys():
				streetData[street1] = [[lat, lng]]
			else:
				streetData[street1].append([lat, lng])

		if not (street2 == '(N/A)' or street2 == None or street2 == '(UNK)'):
			if street2 not in streetData.keys():
				streetData[street2] = [[lat, lng]]
			else:
				streetData[street2].append([lat, lng])

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Portland/bicycle_accidents/final_data_new.json', 'w')
	finalJSON = {'city':'Portland', 'accidents':accidents_with_streetnames, 'streets':streetData, 'clusters':clusters_data}
	json.dump(finalJSON, outputfile)


def portland_street_controller():
	
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Portland/bicycle_accidents/'
	outCSV = open(basePath+'/stretnames2.csv', 'a+')
	jsonData = json.load(open(basePath + 'final_data2.json', 'r'))
	accidents = jsonData['accidents']
	i = 51
	while (i < len(accidents)):
		acc  = accidents[i]
		lat = acc['lat']
		lng = acc['lng']

		[street1_name, street2_name] = get_street_name(lat, lng)

		outCSV.write(str(acc['image_id']) +','+street1_name + ',' +street2_name+'\n')
		print str(i) + ' out of ' + str(len(accidents))
		i+=1
		time.sleep(0.5)


def capitalize_name(street):
	names = street.split()
	newname = ''
	for name in names:
		newname += name.capitalize()
		newname += ' '
	return newname.rstrip()

def split_street(street):
	street_endings = ['AV', 'ST', 'DR', 'BL']
	end = street[-2:]
	if end in street_endings:
		new_name = street[:-2] + ' ' + street[-2:]
	else:
		new_name = str(street)

	#remove u' encoding 

	if new_name[1] == "'":
		new_name = new_name[2:]
	return new_name


def capitalize_streets(city):
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/'+str(city)+'/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data_2013.json','r'))
	accidents = jsonData['accidents']
	streets = jsonData['streets']
	i = 0
	while(i<len(accidents)):
		accident = accidents[i]
		street1 =accident['street1']
		street2 = accident['street2']
		# if street2[0] == "'":
		# 	street2 = street2[1:]

		# accident['street1'] = capitalize_name(split_street(street1))
		# accident['street2'] = capitalize_name(split_street(street2))

		accident['street1'] = capitalize_name(street1)
		accident['street2'] = capitalize_name(street2)
		i+=1
	print i

	updatedStreets = {}
	for street in streets:
		#print street
		updatedStreets[str(capitalize_name(street))] = streets[street]


	# st1st = updatedStreets['2012']['1 Avenue']
	# st11th = updatedStreets['2012']['11 Avenue']
	# print len(st1st), len(st11th)
	# s1D = {}
	# s11D = {}
	# new1st = []
	# for pt1 in st1st:
	# 	s1D[(pt1[0], pt1[1])] = 1
	# for pt11 in st11th:
	# 	s11D[(pt11[0], pt11[1])] = 1

	# new1D = {}
	# for pt in s1D.keys():
	# 	if pt not in s11D.keys():
	# 		new1st.append([pt[0], pt[1]])
	# 		new1D[(pt[0], pt[1])] = 1



	# #print accidents
	# print len(st1st), len(st11th), len(new1st)
	# updatedStreets['2012']['1 Avenue'] = new1st
	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/'+str(city)+'/bicycle_accidents/final_final_data_2013.json', 'w')
	finalJSON = {'city':str(city), 'accidents':accidents, 'streets':updatedStreets, 'clusters':jsonData['clusters']}
	json.dump(finalJSON, outputfile)

def portland_bikeways():
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Portland/bicycle_accidents/'
	geojsonData = geojson.load(open(basePath + 'bikeways.geojson','r'))
	features = geojsonData['features']
	new_features = []
	i = 0
	for feature in features:
		old_coordinates= feature["geometry"]["coordinates"]
		new_coordinates= []
		for coord in old_coordinates:
			new_coordinates.append([coord[0], coord[1]])
		newLineString = geojson.LineString(new_coordinates)
		new_feature = geojson.Feature(geometry=newLineString, properties={'name':'Portland bikeway'}, id=i)
		new_features.append(new_feature)
		i+=1

	featureColleciton = geojson.FeatureCollection(new_features)
	geojson_data = geojson.dumps(featureColleciton)
	output_file = open(basePath + 'bikeways2.geojson','w')
	output_file.write(geojson_data)


def geocode_point(street1, street2, city, state):
	url = 'https://maps.googleapis.com/maps/api/geocode/xml?address='
	url += str(street1)
	url += 'and'
	url += str(street2)
	url += '+'+str(city)+',+'+str(state)+'&sensor=false'

	url = url+'&key=AIzaSyBZnvqy9HEpG-LAQwm_AxDOegMciI9jgP4'
	#secureURL = utils.createSecureURL(url)
	print url
	
	xml = requests.get(url)
	[lat,lng] = getLatLng(xml.content)
	if lat != None and lng != None:
		return [lat, lng]
	else:
		print 'error'
		return [lat, lng]

def getLatLng(response):
	parse = et.XML(response)
	if (parse.find('status').text == 'OK'):
		if (parse.find("result").find("type").text =='intersection'):
			location = parse.find("result").find("geometry").find("location")
			lat = float(location.find("lat").text)
			lng = float(location.find("lng").text)
			return [lat,lng]
		else:
			return [None,None]
	else:
		return [None,None]

def austin_parse_data():
	# filename = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Austin/bicycle_accidents/raw-data.xlsx'
	# workbook = xlreader.load_workbook(filename)
	# sheet = workbook.get_sheet_by_name('Pranav Ramkrishnan 3-21-14')


	accidents = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Austin/bicycle_accidents/all_accidents.csv', 'r')
	accidents = accidents.readlines()
	count = 0
	query_count= 0


	#output = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Austin/bicycle_accidents/all_accidents.csv','w')

	accidentsJSON = []
	streetsJSON = {}

	street_corrections={}
	street_corrections['RED'] = 'RED RIVER ST'
	street_corrections['IH0035'] = 'IH35'
	street_corrections['CHICAN ST'] = 'CHICON ST'
	street_corrections['PEPPIN LN'] = 'PEPPER'
	street_corrections['SHOAL'] = 'SHOAL CREEK BL'
	street_corrections['COMMWEALTH WAY'] = 'COMMONWEALTH'
	street_corrections['WOODLAND DR'] = 'WOODLAND'
	street_corrections['GLADALLPE ST'] = 'GUADALUPE'
	street_corrections['US0290'] = 'BEN WHITE'
	street_corrections['PENTER LN'] = 'PEWTER LN'
	while(count < len(accidents)):
		row = accidents[count].split(',')
		year = int(row[0])
		lat = float(row[2])
		lng = float(row[3])
		street1 = row[4]
		street2 = row[5][:-1]
		severity = row[1]

		if (lat==0 or lng==0):
			
			[lat_qry, lng_qry] = geocode_point(street1, street2, 'Austin', 'TX')
			if lat_qry!=None:
				lat = lat_qry
				lng = lng_qry
			else:
				print street1, street2
				query_count+=1
			time.sleep(0.5)
		
		accObj = {'year':year, 'lat':lat, 'lng':lng, 'severity':severity, 'street1':street1, 'street2':street2}

		if street1 in streetsJSON.keys():
			streetsJSON[street1].append([lat, lng])
		else:
			streetsJSON[street1] = [[lat, lng]]

		if street2 in streetsJSON.keys():
			streetsJSON[street2].append([lat, lng])
		else:
			streetsJSON[street2] = [[lat, lng]]

		accidentsJSON.append(accObj)

		count +=1
		#print count, street1, street2, lat, lng
		#output.write(str(year) + ',' + str(severity) + ',' + str(lat) + ',' + str(lng) + ',' + str(street1) + ',' + str(street2) + '\n')
	
	#print error_list
	print query_count


	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Austin/bicycle_accidents/final_data_new.json', 'w')
	finalJSON= {'city': 'Austin', 'accidents':accidentsJSON, 'streets': streetsJSON}
	json.dump(finalJSON, outputfile)


def manhattan_correction():
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Manhattan/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data2.json','r'))
	csvFile = open(basePath + 'temp_accidents.csv', 'a+')
	accidents = jsonData['accidents']
	[targetLat, targetLng] = [40.753730, -73.985367]
	count=0
	while (count<len(accidents)):
		acc = accidents[count]
		if acc['lat'] != None and acc['lng'] != None:
			if float(acc['lat'] < targetLat):
				print acc['lat'], acc['lng']
				[street1, street2] = get_street_name(float(acc['lat']),float(acc['lng']))
				print count, street1, street2
				acc['street1'] = street1
				acc['street2'] = street2
				csvFile.write(str(acc['lat'])+','+str(acc['lng'])+','+str(street1) + ',' + str(street2) + '\n')
		count+=1
		time.sleep(0.5)

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Manhattan/bicycle_accidents/final_data_corrected.json', 'w')
	finalJSON= {'city': 'Manhattan', 'accidents':accidents, 'streets': jsonData['streets'], 'clusters':jsonData['clusters']}
	json.dump(finalJSON, outputfile)

def manhattan_merge():
	corrected= {}
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Manhattan/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data2.json','r'))
	csvFile = open(basePath + 'temp_accidents.csv', 'r').readlines()
	for pt in csvFile:
		pt = pt.split(',')
		corrected[(float(pt[0]), float(pt[1]))] = [pt[2], pt[3]]

	accidents = jsonData['accidents']
	for acc in accidents:
		if acc['lat'] != None and acc['lng'] != None:
			latlng = (float(acc['lat']), float(acc['lng']))
			if latlng in corrected:
				print 'updated'
				acc['street1'] = corrected[latlng][0]
				acc['street2'] = corrected[latlng][1]

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Manhattan/bicycle_accidents/final_data_corrected.json', 'w')
	finalJSON= {'city': 'Manhattan', 'accidents':accidents, 'streets': jsonData['streets'], 'clusters':jsonData['clusters']}
	json.dump(finalJSON, outputfile)

def manhattan_correction2():
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Manhattan/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data3.json','r'))
	accidents = jsonData['accidents']
	for acc in accidents:
		street1 = acc['street1']
		street2 = acc['street2']
		if street1 == 'West 45 Street' or street2 == 'West 45 Street':
			if street1 =='6 Avenue' or street2 == '6 Avenue':
				acc['street2'] = 'York Street'

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Manhattan/bicycle_accidents/final_data5.json', 'w')
	finalJSON= {'city': 'Manhattan', 'accidents':accidents, 'streets': jsonData['streets'], 'clusters':jsonData['clusters']}
	json.dump(finalJSON, outputfile)


def cambridge_correction():
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data3.json','r'))
	accidents = jsonData['accidents']
	all2012 = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/error_all_2012.csv','w')
	accidents_no_dups = []
	counts = {'2010':0, '2011':0, '2012':0, '2013':0}
	for acc in accidents:
		counts[acc['year']]+=1
		line = ''
		if acc['year'] == '2012':
			line+=acc['year']
			line+= ','
			if acc['street1'] != "(n/a)":
				line+=acc['street1']
				line+= ','
			if acc['street2'] != "(n/a)":
				line+=acc['street2']
				

			print line
			all2012.write(line+'\n')
			accidents_no_dups.append(line)	

	print len(accidents_no_dups)
	print len(set(accidents_no_dups))
	print counts

def cambridge_2012():
	filename = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/Cambridge 2010 to 2013 all Accident Data.csv'
	points = open(filename,'r').readlines()[0].split('\r')
	only2012 = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/only_2012.csv','w')
	for pt in points:

		pt2 = pt.split(',')
		date = pt2[0].split('/')
		year = date[-1]
		type1 = pt2[3]
		type2 = pt2[4]
		#print date,year
		if type1 =='Bicycle' or type2 =='Bicycle':
			line = ''
			if '2012' in year:
				street1 = capitalize_name(pt2[-1])
				street2 = capitalize_name(pt2[-2])
				line+=year
				line+= ','
				if street1 != '':
					line+=street1
					line+= ','
				if street2 != '':
					line+=street2
				print line
				only2012.write(line +'\n')

def cambridge_crosscheck():
	correct = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/only_2012.csv','r')
	#errors = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/error_all_2012.csv','r')
	correct_points = correct.readlines()
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_data3.json','r'))
	accidents = jsonData['accidents']
	corrected_accidents = []
	acc_2012 = []
	for ac in accidents:
		if ac['year'] != '2012':
			corrected_accidents.append(ac)
		else:
			acc_2012.append(ac)

	print len(acc_2012)

	
	i = 0
	check_count=0
	while(i<len(correct_points)):
		correct_line = correct_points[i][:-1]
		acc = acc_2012[i]
		line = ''
		if acc['year'] == '2012':
			line+=acc['year']
			line+= ','
			if acc['street2'] != "(n/a)":
				line+=acc['street2']
				line+= ','
			if acc['street1'] != "(n/a)":
				line+=acc['street1']

		if correct_line == line:
			print "check"
			check_count+=1
			corrected_accidents.append(acc)
		else:
			print correct_line, line
		i+=1

	print check_count
	print len(corrected_accidents)
	streetsJSON = {}

	for acc in corrected_accidents:
		print acc
		street1 = acc['street1']
		street2 = acc['street2']
		lat = acc['lat']
		lng = acc['lng']


		if street1 in streetsJSON.keys() and street1 !='(n/a)':
			streetsJSON[street1].append([lat, lng])
		else:
			streetsJSON[street1] = [[lat, lng]]

		if street2 in streetsJSON.keys() and street2 !='(n/a)':
			streetsJSON[street2].append([lat, lng])
		else:
			streetsJSON[street2] = [[lat, lng]]

	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/Cambridge/bicycle_accidents/final_corrected.json', 'w')
	finalJSON= {'city': 'Cambridge', 'accidents':corrected_accidents, 'streets': streetsJSON}
	json.dump(finalJSON, outputfile)


def cleanJsons(city,url):
	basePath = '/Users/pranavramkrishnan/Desktop/Research/maps/static/data/' + str(city) + '/bicycle_accidents/'
	jsonData = json.load(open(basePath + 'final_final_data_2013.json','r'))
	accidents = jsonData['accidents']
	outputfile = open('/Users/pranavramkrishnan/Desktop/Research/maps/static/data/'+str(city)+'/bicycle_accidents/final_prettified.json', 'w')
	finalJSON= {'city': str(city), 'source':url, 'accidents':accidents}
	json.dump(finalJSON, outputfile)




#cleanJsons('Cambridge','http://www.cambridgema.gov/cpd/')


## Run commands in the following orer ##
#makeStreetClusters('Manhattan')
downloadStreetViews('Manhattan',1901)
capitalize_streets('Manhattan')
cleanJsons('Manhattan','https://nycopendata.socrata.com/')

##########




#cambridge_crosscheck()
#cambridge_2012()
#cambridge_correction()
#manhattan_merge()
#manhattan_correction2()
#downloadStreetViews('Brooklyn')
#austin_parse_data()
#portland_bikeways()
#portland_street()
#capitalize_streets('Manhattan')
#makeStreetClusters('Brooklyn')
#mergeJSONdata('Portland')
#readXLSX('Chicago')